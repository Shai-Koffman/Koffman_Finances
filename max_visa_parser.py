#!/usr/bin/env python
# -*- coding: utf-8 -*-

from dateutil.parser import parse
import openpyxl
from transactions import VisaMaxTransaction
from currency import CoinType
from categories import get_category


def parse_visa_max_file(filename):
    wb = openpyxl.load_workbook(filename)
    items = []
    for sheet in wb.worksheets:
        sheet_title = sheet.title
        number_of_rows = sheet.max_row
        for row in range(2, number_of_rows + 1):  # Adjusted to start from row 2 assuming row 1 has headers

            #company name
            company = sheet.cell(row, 2).value 
            known_category = get_category(company)
            #visa_max_category 
            visa_max_category = sheet.cell(row, 3).value
            
            #visa card number
            visa_card_number = sheet.cell(row, 4).value

            expense_sum:int = int(sheet.cell(row, 6).value)

            if expense_sum < 0:
                gain_sum = -expense_sum
                expense_sum = 0
            else:
                gain_sum = 0
            coin:CoinType = CoinType.get_coin_from_symbol(sheet.cell(row, 7).value)

            #expense date
            expense_date_cell_value = sheet.cell(row, 10).value
            parsedDate = parse(expense_date_cell_value, dayfirst=True)  
            
            item = VisaMaxTransaction(sheet_title=sheet_title,
                                company=company,
                                known_category=known_category,
                                visa_max_category=visa_max_category,
                                visa_card_number=visa_card_number,
                                expense_sum=expense_sum,
                                gain_sum=gain_sum,
                                coin=coin,
                                expense_date=parsedDate)
            items.append(item)
    return items



