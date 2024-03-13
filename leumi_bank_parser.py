from openpyxl import load_workbook
import datetime
from expenses import BankExpense  

def parse_bank_file(filename):
    wb = load_workbook(filename)
    items = []
    for sheet in wb.worksheets:
        
        number_of_rows = sheet.max_row
        for row in range(1, number_of_rows + 1):
            parsedDate = datetime.datetime.fromordinal(
                datetime.datetime(1900, 1, 1).toordinal() + int(sheet.cell(row, 0).value) - 2)
            company = sheet.cell(row, 1).value
            in_file_expense = sheet.cell(row, 3).value
            if not in_file_expense:
                continue
            expense = float(in_file_expense)
            item = BankExpense(parsedDate, company, expense)
            items.append(item)
        # print "number of items in %s is %d" % (filename, len(items))
        return items
